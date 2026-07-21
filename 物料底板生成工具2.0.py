# -*- coding: utf-8 -*-
"""
物料底板生成工具 2.0
功能：
  Tab1 - 问卷+商城生成（各自独立行，不合并）
  Tab2 - 三合一提报（主动提报 / 费率换签 / 高端物料）
模板：38列 = 8基础信息 + 24物料 + 6尾部
"""

import os
import re
import sys
import traceback
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except Exception:
    HAS_DND = False


# ============================================================
# 输出目录
# ============================================================
def get_output_dir():
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    out_dir = os.path.join(desktop, "Catdesk file")
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


# ============================================================
# ★★★ 物料列配置区 — 后续修改物料列只需改这里 ★★★
# ============================================================
MATERIAL_COLUMNS = [
    "大膠袋", "小膠袋",
    "單杯無紡布袋", "雙杯無紡布袋", "3號無紡布袋", "4號無紡布袋",
    "紙袋", "打印紙", "紙漿杯2托", "封簽",
    "玻璃門貼", "優惠貼", "海報（中文版）", "亞克力水牌",
    "pos機", "充電頭", "充電線",
    "紙漿杯托-4托", "飲品封口紙", "雙杯裝紙袋", "單杯裝紙袋",
    "高端纸袋（黑色普通）", "高端保溫袋", "高端封簽",
]

# ============================================================
# 模板列定义（= 基础信息 + 物料列 + 尾部信息）共38列
# ============================================================
INFO_COLUMNS = [
    "提交日期", "訂單類型", "下單時間", "售賣/贈送",
    "商家ID", "商家名稱", "收貨電話", "收貨地址",
]
TAIL_COLUMNS = [
    "備註", "商家營業時間", "調賬費用", "發起調賬", "總數量", "物流",
]
TEMPLATE_COLUMNS = INFO_COLUMNS + MATERIAL_COLUMNS + TAIL_COLUMNS

# ============================================================
# 问卷 CSP 映射（问卷列名 → 模板物料列名）
# ============================================================
CSP_MAPPINGS = [
    {"templateName": "大膠袋",       "cnPattern": "11.膠袋",    "enPattern": "22.Plastic Bag"},
    {"templateName": "小膠袋",       "cnPattern": "12.小膠袋",   "enPattern": "23.Small Plastic Bag"},
    {"templateName": "單杯無紡布袋", "cnPattern": "6.無紡布袋（保溫） | HK$150", "enPattern": "17.Non-woven bag(single serve cup"},
    {"templateName": "雙杯無紡布袋", "cnPattern": "7.無紡布袋（保溫）| HK$180", "enPattern": "18.Non-woven bag(double serve cup"},
    {"templateName": "3號無紡布袋",  "cnPattern": "8.無紡布袋（保溫） | HK$230", "enPattern": "19.Non-woven bag( medium size"},
    {"templateName": "4號無紡布袋",  "cnPattern": "9.無紡布袋（保溫）| HK$280", "enPattern": "20.Non-woven bag( Standard size"},
    {"templateName": "紙袋",         "cnPattern": "5.紙袋",      "enPattern": "16.Paper Bag"},
    {"templateName": "打印紙",       "cnPattern": "14.打印紙",   "enPattern": "25.Print Roll"},
    {"templateName": "紙漿杯2托",    "cnPattern": "10.紙漿杯托", "enPattern": "21.Cup Holder"},
    {"templateName": "封簽",         "cnPattern": "13.食品安全封簽", "enPattern": "24.Food Safety Seal"},
]

# ============================================================
# 商家服务市场订单 CSP 映射（规格 → 模板物料列名）
# ============================================================
SERVICE_CSP_MAPPINGS = [
    {"specName": "大膠袋",         "templateName": "大膠袋"},
    {"specName": "小膠袋",         "templateName": "小膠袋"},
    {"specName": "規格一單杯裝",   "templateName": "單杯無紡布袋"},
    {"specName": "規格二雙杯裝",   "templateName": "雙杯無紡布袋"},
    {"specName": "規格三中號袋",   "templateName": "3號無紡布袋"},
    {"specName": "規格四大號袋",   "templateName": "4號無紡布袋"},
    {"specName": "通版紙袋",       "templateName": "紙袋"},
    {"specName": "熱敏紙",         "templateName": "打印紙"},
    {"specName": "紙漿杯托-2杯裝", "templateName": "紙漿杯2托"},
    {"specName": "紙漿杯托-4杯裝", "templateName": "紙漿杯托-4托"},
    {"specName": "食物安全封簽",   "templateName": "封簽"},
    {"specName": "食品級飲品封口紙", "templateName": "飲品封口紙"},
]

CAP_LIMIT = 10

# ============================================================
# 三合一提报映射配置
# ============================================================
# 主动提报映射（主动提报物料名 → 底板物料列名）
ACTIVE_REPORT_MAPPINGS = {
    "無紡布袋4號": "4號無紡布袋",
    "無紡布袋3號": "3號無紡布袋",
    "無紡布袋2號": "雙杯無紡布袋",
    "無紡布袋1號": "單杯無紡布袋",
    "紙袋": "紙袋",
    "大膠袋": "大膠袋",
    "小膠袋": "小膠袋",
    "雙杯杯托": "紙漿杯2托",
    "四杯杯托": "紙漿杯托-4托",
}

# 费率换签映射（费率换签物料类型 → 底板物料列名）
RATE_SIGN_MAPPINGS = {
    "紙袋": "紙袋",
    "无纺布袋-1号": "單杯無紡布袋",
    "无纺布袋-2号": "雙杯無紡布袋",
    "无纺布袋-3号": "3號無紡布袋",
    "无纺布袋-4号": "4號無紡布袋",
    "塑胶袋-大": "大膠袋",
    "塑胶袋-小": "小膠袋",
}

# 高端物料映射（高端物料名 → 底板物料列名）
PREMIUM_MAPPINGS = {
    "高端保溫袋（黑色）": "高端保溫袋",
    "高端紙袋（黑色）": "高端纸袋（黑色普通）",
    "高端封簽（黑色）": "高端封簽",
    "4號保溫袋（黃色）": "4號無紡布袋",
    "3號保溫袋（黃色）": "3號無紡布袋",
    "2號保溫袋（黃色）": "雙杯無紡布袋",
    "1號保溫袋（黃色）": "單杯無紡布袋",
    "四格杯托": "紙漿杯托-4托",
    "兩格杯托": "紙漿杯2托",
    "大膠袋（黃色）": "大膠袋",
    "小膠袋（黃色）": "小膠袋",
    "通版紙袋（黃色）": "紙袋",
}

# 高端物料忽略列表
PREMIUM_IGNORE = {"Laisun紙袋（For Ken）", "Laisun紙袋(For Ken)"}

# Excel 日期序列号基准
EXCEL_EPOCH = datetime(1899, 12, 30)


# ============================================================
# 工具函数
# ============================================================
def extract_quantity(val):
    """取字符串开头的数字，例如 '3-已选' -> 3"""
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return None
    m = re.match(r"^\s*(\d+)\s*", s)
    return int(m.group(1)) if m else None


def to_num(val):
    if val is None or val == "":
        return None
    try:
        if isinstance(val, (int, float)):
            return val
        s = str(val).strip()
        m = re.match(r"^\s*(\d+)", s)
        return int(m.group(1)) if m else None
    except Exception:
        return None


def read_excel_all_str(path):
    """读取 Excel，全部作为字符串/原始值处理，第一行为表头。"""
    df = pd.read_excel(path, header=None, dtype=object)
    df = df.fillna("")
    return df.values.tolist()


def find_col_index(headers, pattern):
    for i, h in enumerate(headers):
        if pattern in str(h):
            return i
    return -1


def today_str():
    return datetime.now().strftime("%Y-%m-%d")


# ============================================================
# 解析问卷 Excel → 行列表
# ============================================================
def parse_questionnaire(path):
    raw = read_excel_all_str(path)
    if len(raw) < 2:
        raise ValueError("问卷文件为空或格式不正确")

    headers = [str(h).strip() for h in raw[0]]
    rows = [r for r in raw[1:] if any(str(c) != "" for c in r)]

    if "shopId" not in headers:
        raise ValueError("找不到 shopId 列，请确认 Excel 格式是否为问卷结果")
    shop_id_idx = headers.index("shopId")
    submit_time_idx = 2

    col_mapping = {}
    mapped_count = 0
    for m in CSP_MAPPINGS:
        cn_idx = find_col_index(headers, m["cnPattern"])
        en_idx = find_col_index(headers, m["enPattern"])
        col_mapping[m["templateName"]] = {"cnIdx": cn_idx, "enIdx": en_idx}
        if cn_idx >= 0 or en_idx >= 0:
            mapped_count += 1

    if mapped_count == 0:
        raise ValueError("无法匹配任何商品列，请确认 Excel 格式是否为问卷结果")

    today = today_str()
    result_rows = []

    for row in rows:
        shop_id = str(row[shop_id_idx]).strip() if shop_id_idx < len(row) else ""
        if not shop_id:
            continue

        new_row = {col: "" for col in TEMPLATE_COLUMNS}
        new_row["提交日期"] = today
        new_row["訂單類型"] = "问卷"
        new_row["下單時間"] = str(row[submit_time_idx]).strip() if 0 <= submit_time_idx < len(row) else ""
        new_row["商家ID"] = shop_id
        new_row["物流"] = "順豐"
        new_row["_source"] = "问卷"

        for m in CSP_MAPPINGS:
            info = col_mapping[m["templateName"]]
            qty = None
            if info["cnIdx"] >= 0 and info["cnIdx"] < len(row) and row[info["cnIdx"]] != "":
                qty = extract_quantity(row[info["cnIdx"]])
            if qty is None and info["enIdx"] >= 0 and info["enIdx"] < len(row) and row[info["enIdx"]] != "":
                qty = extract_quantity(row[info["enIdx"]])
            if qty is not None and qty > 0:
                new_row[m["templateName"]] = qty

        result_rows.append(new_row)

    return result_rows


# ============================================================
# 解析商家服务市场订单 Excel → 行列表
# ============================================================
def parse_service_order(path):
    raw = read_excel_all_str(path)
    if len(raw) < 2:
        raise ValueError("服务订单文件为空或格式不正确")

    header_row_idx = 0
    headers = [str(h).strip() for h in raw[0]]
    if len(raw) > 1:
        first_data_row = [str(h).strip() for h in raw[1]]
        if any(k in first_data_row for k in ("訂單號", "订单号", "Keeta門店ID", "Keeta门店ID")):
            header_row_idx = 1
            headers = first_data_row

    data_start_idx = header_row_idx + 1
    rows = [r for r in raw[data_start_idx:] if any(str(c) != "" for c in r)]

    def idx_of(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return -1

    shop_id_col = idx_of("Keeta门店ID", "Keeta門店ID")
    order_time_col = idx_of("下单时间", "下單時間")
    spec_col = idx_of("规格", "規格")
    qty_col = idx_of("数量", "數量")

    if shop_id_col == -1:
        raise ValueError("服务订单找不到 Keeta门店ID 列")

    # 按门店聚合（同门店同规格数量累加）
    service_map = {}
    spec_match_count = {m["specName"]: 0 for m in SERVICE_CSP_MAPPINGS}
    unmapped_sv = set()

    for row in rows:
        shop_id = str(row[shop_id_col]).strip() if shop_id_col < len(row) else ""
        if not shop_id or shop_id in ("Keeta门店ID", "Keeta門店ID"):
            continue

        if shop_id not in service_map:
            service_map[shop_id] = {"orderTime": "", "items": {}}

        if 0 <= order_time_col < len(row) and row[order_time_col]:
            new_time = str(row[order_time_col]).strip()
            if not service_map[shop_id]["orderTime"] or new_time > service_map[shop_id]["orderTime"]:
                service_map[shop_id]["orderTime"] = new_time

        spec_val = str(row[spec_col]).strip() if 0 <= spec_col < len(row) else ""
        if not spec_val or spec_val in ("规格", "規格"):
            continue

        matched = False
        for m in SERVICE_CSP_MAPPINGS:
            if spec_val == m["specName"]:
                raw_qty = row[qty_col] if 0 <= qty_col < len(row) else 0
                qty = to_num(raw_qty) or 0
                if qty > 0:
                    items = service_map[shop_id]["items"]
                    items[m["templateName"]] = items.get(m["templateName"], 0) + qty
                    spec_match_count[m["specName"]] += 1
                matched = True
                break
        if not matched:
            unmapped_sv.add(spec_val)

    # 转为行列表
    today = today_str()
    result_rows = []
    for shop_id, info in service_map.items():
        new_row = {col: "" for col in TEMPLATE_COLUMNS}
        new_row["提交日期"] = today
        new_row["訂單類型"] = "商城"
        new_row["下單時間"] = info["orderTime"]
        new_row["商家ID"] = shop_id
        new_row["物流"] = "順豐"
        new_row["_source"] = "商城"
        for tmpl_name, qty in info["items"].items():
            new_row[tmpl_name] = qty
        result_rows.append(new_row)

    return result_rows, spec_match_count, sorted(unmapped_sv)


# ============================================================
# 解析门店资料 Vlookup Excel
# ============================================================
def parse_vlookup(path):
    raw = read_excel_all_str(path)
    if len(raw) < 2:
        raise ValueError("门店资料文件为空或格式不正确")

    headers = [str(h).strip() for h in raw[0]]
    shop_id_col = contact_col = address_col = shop_name_col = -1
    for i, h in enumerate(headers):
        if h in ("门店id", "門店id", "门店ID", "門店ID", "商家ID", "商家id"):
            shop_id_col = i
        if h in ("contact", "電話", "电话", "收貨電話", "收货电话"):
            contact_col = i
        if h in ("address", "地址", "收貨地址", "收货地址"):
            address_col = i
        if h in ("门店名称", "門店名稱", "门店名", "門店名", "商家名稱", "商家名称"):
            shop_name_col = i

    if shop_id_col == -1:
        raise ValueError("门店资料找不到 门店ID/商家ID 列")

    vl_map = {}
    for row in raw[1:]:
        if shop_id_col >= len(row) or not row[shop_id_col]:
            continue
        shop_id = str(row[shop_id_col]).strip()
        if not shop_id:
            continue
        vl_map[shop_id] = {
            "shopName": str(row[shop_name_col]).strip() if 0 <= shop_name_col < len(row) else "",
            "contact": str(row[contact_col]).strip() if 0 <= contact_col < len(row) else "",
            "address": str(row[address_col]).strip() if 0 <= address_col < len(row) else "",
        }
    return vl_map


# ============================================================
# 应用 Vlookup
# ============================================================
def apply_vlookup(row, vl_data):
    if not vl_data:
        return
    shop_id = row.get("商家ID", "")
    if shop_id and shop_id in vl_data:
        info = vl_data[shop_id]
        if info["shopName"]:
            row["商家名稱"] = info["shopName"]
        if info["contact"]:
            row["收貨電話"] = info["contact"]
        if info["address"]:
            row["收貨地址"] = info["address"]


# ============================================================
# 数量上限检查
# ============================================================
def cap_check(rows):
    cap_log = []
    for row in rows:
        capped = []
        for col in MATERIAL_COLUMNS:
            val = row.get(col, "")
            if val == "" or val is None:
                continue
            num = val if isinstance(val, (int, float)) else to_num(val)
            if num is None or num <= 0:
                continue
            if num > CAP_LIMIT:
                row[col] = CAP_LIMIT
                capped.append(f"{col}({int(num)}→{CAP_LIMIT})")
        if capped:
            cap_log.append({
                "shopId": row.get("商家ID", ""),
                "orderType": row.get("訂單類型", ""),
                "details": capped,
            })
    return cap_log


# ============================================================
# 统计
# ============================================================
def count_items(rows):
    count = 0
    for row in rows:
        for col in MATERIAL_COLUMNS:
            v = row.get(col)
            if v not in ("", None):
                num = v if isinstance(v, (int, float)) else to_num(v)
                if num and num > 0:
                    count += 1
    return count


def count_total(rows):
    total = 0
    for row in rows:
        for col in MATERIAL_COLUMNS:
            v = row.get(col)
            if v not in ("", None):
                num = v if isinstance(v, (int, float)) else to_num(v)
                if num and num > 0:
                    total += num
    return total


# ============================================================
# 三合一提报解析函数
# ============================================================
def excel_serial_to_date(serial):
    """Excel 日期序列号转日期字符串"""
    try:
        n = int(float(serial))
        return (EXCEL_EPOCH + timedelta(days=n)).strftime("%Y-%m-%d")
    except Exception:
        return str(serial).strip()


def normalize_date(val):
    """将各种日期格式统一为 YYYY-MM-DD 字符串"""
    if val is None or val == "":
        return ""
    from datetime import datetime as dt
    if isinstance(val, dt):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if s.lower() == "nan":
        return ""
    # Excel 序列号
    if re.match(r"^\d{4,6}(\.\d+)?$", s) and not re.match(r"^\d{4}-", s):
        return excel_serial_to_date(s)
    # 带时间的日期字符串，截取日期部分
    if " " in s:
        return s.split(" ")[0]
    return s


def parse_active_report(path):
    """解析主动提报模板。按列位置读取：C=shopid, G=物料名, H=数量, A=创建时间, J=备注"""
    raw = read_excel_all_str(path)
    if len(raw) < 2:
        raise ValueError("主动提报文件为空或格式不正确")

    data_rows = raw[1:]  # 第一行为表头，跳过
    today = today_str()
    result_rows = []
    warnings = []  # 未映射物料
    cap_warnings = []  # 超过10的提醒

    for row in data_rows:
        if len(row) < 8:
            continue
        shop_id = str(row[2]).strip()  # C列
        if not shop_id or shop_id == "":
            continue

        material_name = str(row[6]).strip() if len(row) > 6 else ""  # G列（位置固定）
        qty_raw = row[7] if len(row) > 7 else 0  # H列
        remark = str(row[9]).strip() if len(row) > 9 and row[9] and str(row[9]).strip() else ""  # J列

        if not material_name:
            continue

        template_col = ACTIVE_REPORT_MAPPINGS.get(material_name)
        if not template_col:
            warnings.append(material_name)
            continue

        qty = to_num(qty_raw) or 0
        if qty <= 0:
            continue
        if qty > CAP_LIMIT:
            qty = CAP_LIMIT
            cap_warnings.append(f"门店 {shop_id} - {material_name}({qty_raw}→{CAP_LIMIT})")

        order_time = normalize_date(row[0])  # A列

        new_row = {col: "" for col in TEMPLATE_COLUMNS}
        new_row["提交日期"] = today
        new_row["訂單類型"] = "主动提报"
        new_row["下單時間"] = order_time
        new_row["商家ID"] = shop_id
        new_row["物流"] = "順豐"
        new_row["備註"] = remark
        new_row[template_col] = qty
        new_row["_source"] = "主动提报"

        result_rows.append(new_row)

    return result_rows, sorted(set(warnings)), cap_warnings


def parse_rate_sign(path):
    """解析费率换签模版。按列名读取：B=门店id, C=门店名称, F=物料类型, J=发出日期"""
    raw = read_excel_all_str(path)
    if len(raw) < 2:
        raise ValueError("费率换签文件为空或格式不正确")

    headers = [str(h).strip() for h in raw[0]]

    def idx_of(*names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h:
                    return i
        return -1

    shop_id_col = idx_of("門店id", "门店id", "門店ID", "门店ID")
    shop_name_col = idx_of("门店名称", "門店名稱")
    material_col = idx_of("物料類型", "物料类型")
    send_date_col = idx_of("發出日期", "发出日期")

    if shop_id_col == -1:
        raise ValueError("费率换签找不到 门店id 列")
    if material_col == -1:
        raise ValueError("费率换签找不到 物料類型 列")

    today = today_str()
    result_rows = []
    warnings = []

    for row in raw[1:]:
        if len(row) <= max(shop_id_col, material_col):
            continue
        shop_id = str(row[shop_id_col]).strip()
        if not shop_id:
            continue

        material_name = str(row[material_col]).strip() if row[material_col] else ""
        if not material_name or material_name.lower() == "nan":
            continue

        template_col = RATE_SIGN_MAPPINGS.get(material_name)
        if not template_col:
            warnings.append(material_name)
            continue

        shop_name = str(row[shop_name_col]).strip() if 0 <= shop_name_col < len(row) and row[shop_name_col] else ""
        send_date = normalize_date(row[send_date_col]) if 0 <= send_date_col < len(row) else ""

        new_row = {col: "" for col in TEMPLATE_COLUMNS}
        new_row["提交日期"] = today
        new_row["訂單類型"] = "费率换签"
        new_row["下單時間"] = send_date
        new_row["商家ID"] = shop_id
        new_row["商家名稱"] = shop_name
        new_row["物流"] = "順豐"
        new_row[template_col] = 1  # 固定数量为1
        new_row["_source"] = "费率换签"

        result_rows.append(new_row)

    return result_rows, sorted(set(warnings))


def parse_premium(path):
    """解析高端物料模版。按列位置：B=shopid, C=门店名称, E=提报物料, F=数量, A=创建时间, G=备注"""
    raw = read_excel_all_str(path)
    if len(raw) < 2:
        raise ValueError("高端物料文件为空或格式不正确")

    today = today_str()
    result_rows = []
    warnings = []  # 未映射物料
    cap_warnings = []  # 超过10的提醒
    out_of_stock = []  # 无货提醒

    for row in raw[1:]:
        if len(row) < 6:
            continue
        shop_id = str(row[1]).strip()  # B列
        if not shop_id or shop_id == "":
            continue
        # 清理 shopid 中的换行符
        shop_id = shop_id.replace("\n", "").strip()

        shop_name = str(row[2]).strip() if row[2] else ""  # C列
        material_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""  # E列
        qty_raw = row[5] if len(row) > 5 else 0  # F列
        remark = str(row[6]).strip() if len(row) > 6 and row[6] and str(row[6]).strip() and str(row[6]).strip().lower() != "nan" else ""  # G列

        if not material_name or material_name.lower() == "nan":
            continue

        # 检查无货
        if "無貨" in material_name or "无货" in material_name:
            out_of_stock.append(f"门店 {shop_id}({shop_name}) - {material_name}")
            continue

        # 检查忽略列表
        skip = False
        for ignore_name in PREMIUM_IGNORE:
            if ignore_name in material_name:
                skip = True
                break
        if skip:
            continue

        # 检查圣诞款
        if "圣诞" in material_name or "聖誕" in material_name:
            continue

        template_col = PREMIUM_MAPPINGS.get(material_name)
        if not template_col:
            warnings.append(material_name)
            continue

        qty = to_num(qty_raw) or 0
        if qty <= 0:
            continue
        if qty > CAP_LIMIT:
            cap_warnings.append(f"门店 {shop_id}({shop_name}) - {material_name}: 数量={int(qty)}（超过{CAP_LIMIT}，请确认）")
            # 不截断，按实际数量填入

        order_time = normalize_date(row[0])  # A列

        new_row = {col: "" for col in TEMPLATE_COLUMNS}
        new_row["提交日期"] = today
        new_row["訂單類型"] = "高端物料"
        new_row["下單時間"] = order_time
        new_row["商家ID"] = shop_id
        new_row["商家名稱"] = shop_name
        new_row["物流"] = "順豐"
        new_row["備註"] = remark
        new_row[template_col] = qty
        new_row["_source"] = "高端物料"

        result_rows.append(new_row)

    return result_rows, sorted(set(warnings)), cap_warnings, out_of_stock


# ============================================================
# 导出 Excel
# ============================================================
HEADER_FILL = PatternFill(start_color="FF6366F1", end_color="FF6366F1", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11, name="Microsoft JhengHei")
COLOR_QS = Font(bold=True, color="FF4F46E5", name="Microsoft JhengHei")    # 蓝：问卷
COLOR_SV = Font(bold=True, color="FFD97706", name="Microsoft JhengHei")    # 橙：商城
COLOR_AR = Font(bold=True, color="FF059669", name="Microsoft JhengHei")    # 绿：主动提报
COLOR_RS = Font(bold=True, color="FF7C3AED", name="Microsoft JhengHei")    # 紫：费率换签
COLOR_PM = Font(bold=True, color="FFDC2626", name="Microsoft JhengHei")    # 红：高端物料
DEFAULT_FONT = Font(size=11, name="Microsoft JhengHei")

COL_WIDTHS = {
    "提交日期": 14, "訂單類型": 10, "下單時間": 20, "售賣/贈送": 10,
    "商家ID": 14, "商家名稱": 30, "收貨電話": 16, "收貨地址": 40,
    "備註": 20, "商家營業時間": 16, "調賬費用": 10, "發起調賬": 10,
    "總數量": 10, "物流": 8,
}


def export_excel(rows, save_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "物料底板"

    # 表头
    ws.append(TEMPLATE_COLUMNS)
    for c_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # 列宽
    for i, col in enumerate(TEMPLATE_COLUMNS, start=1):
        letter = get_column_letter(i)
        if col in COL_WIDTHS:
            ws.column_dimensions[letter].width = COL_WIDTHS[col]
        elif col in MATERIAL_COLUMNS:
            ws.column_dimensions[letter].width = 14 if len(col) > 4 else 10
        else:
            ws.column_dimensions[letter].width = 14

    # 数据行
    col_idx_map = {col: i + 1 for i, col in enumerate(TEMPLATE_COLUMNS)}
    for row in rows:
        row_values = []
        for col in TEMPLATE_COLUMNS:
            val = row.get(col, "")
            if val == "" or val is None:
                row_values.append(None)
            elif col in ("提交日期", "訂單類型", "下單時間", "商家ID"):
                row_values.append(str(val))
            else:
                num = val if isinstance(val, (int, float)) else to_num(val)
                row_values.append(num if num is not None else str(val))
        ws.append(row_values)

    # 样式：物料列居中 + 来源颜色
    for r_idx in range(2, ws.max_row + 1):
        source = rows[r_idx - 2].get("_source", "")
        if source == "问卷":
            color_font = COLOR_QS
        elif source == "商城":
            color_font = COLOR_SV
        elif source == "主动提报":
            color_font = COLOR_AR
        elif source == "费率换签":
            color_font = COLOR_RS
        elif source == "高端物料":
            color_font = COLOR_PM
        else:
            color_font = DEFAULT_FONT

        for col in MATERIAL_COLUMNS:
            c_idx = col_idx_map[col]
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = Alignment(horizontal="center")
            if cell.value is not None:
                cell.font = color_font

        for base_col in ("提交日期", "訂單類型", "下單時間", "商家ID"):
            c_idx = col_idx_map[base_col]
            ws.cell(row=r_idx, column=c_idx).font = DEFAULT_FONT

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    last_col_letter = get_column_letter(len(TEMPLATE_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    # 表头底部边框
    thick_bottom = Border(bottom=Side(style="medium", color="FF4338CA"))
    for c_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
        ws.cell(row=1, column=c_idx).border = thick_bottom

    wb.save(save_path)


# ============================================================
# GUI
# ============================================================
class MaterialToolApp:
    def __init__(self, root):
        self.root = root
        self.root.title("物料底板生成工具 2.0")
        self.root.geometry("760x700")
        self.root.configure(bg="#f5f7fa")
        self.root.resizable(True, True)
        self.root.minsize(760, 500)

        self.qs_data = None
        self.sv_data = None
        self.sv_spec_count = None
        self.vl_data = None
        self.merged_rows = None
        self.unmapped_sv = []

        self._build_ui()

    def _build_ui(self):
        title = tk.Label(self.root, text="物料底板生成工具 2.0", font=("Microsoft JhengHei", 20, "bold"),
                         bg="#f5f7fa", fg="#1e293b")
        title.pack(pady=(20, 4))
        subtitle = tk.Label(self.root, text="问卷+商城各自独立行，不合并", font=("Microsoft JhengHei", 11),
                            bg="#f5f7fa", fg="#64748b")
        subtitle.pack(pady=(0, 12))

        # ttk 样式
        style = ttk.Style()
        style.configure("TNotebook", background="#f5f7fa", borderwidth=0)
        style.configure("TNotebook.Tab", font=("Microsoft JhengHei", 11), padding=(16, 6))
        style.map("TNotebook.Tab",
                  background=[("selected", "#6366f1")],
                  foreground=[("selected", "#ffffff")])

        notebook = ttk.Notebook(self.root)
        notebook.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        self.tab1 = tk.Frame(notebook, bg="#f5f7fa")
        self.tab2 = tk.Frame(notebook, bg="#f5f7fa")
        notebook.add(self.tab1, text="📋  问卷+商城生成")
        notebook.add(self.tab2, text="🔄  三合一提报")

        self._build_tab1()
        self._build_tab2()

    # ---------- 可滚动容器 ----------
    def _make_scrollable(self, parent):
        canvas = tk.Canvas(parent, bg="#f5f7fa", highlightthickness=0)
        scrollbar = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg="#f5f7fa")

        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        return inner

    # ---------- Tab1：问卷+商城 ----------
    def _build_tab1(self):
        inner = self._make_scrollable(self.tab1)

        zones_frame = tk.Frame(inner, bg="#f5f7fa")
        zones_frame.pack(pady=14)

        self.qs_zone = self._make_drop_zone(zones_frame, "问卷结果", "#6366f1", 0, self.on_qs_selected)
        self.sv_zone = self._make_drop_zone(zones_frame, "商家服务订单", "#d97706", 1, self.on_sv_selected)
        self.vl_zone = self._make_drop_zone(zones_frame, "门店资料 Vlookup", "#16a34a", 2, self.on_vl_selected)

        self.status_label = tk.Label(
            inner, text="请上传「问卷结果」或「商家服务订单」以开始",
            font=("Microsoft JhengHei", 11), bg="#f5f7fa", fg="#334155",
            wraplength=700, justify="center")
        self.status_label.pack(pady=(14, 8))

        stats_frame = tk.Frame(inner, bg="#f5f7fa")
        stats_frame.pack(pady=4)
        self.stat_qs = self._make_stat_card(stats_frame, "问卷门店数", 0)
        self.stat_sv = self._make_stat_card(stats_frame, "商城门店数", 1)
        self.stat_items = self._make_stat_card(stats_frame, "有下单品项", 2)
        self.stat_total = self._make_stat_card(stats_frame, "总下单数", 3)

        self.cap_label = tk.Label(
            inner, text="", font=("Microsoft JhengHei", 9), bg="#f5f7fa",
            fg="#dc2626", wraplength=700, justify="left")
        self.cap_label.pack(pady=(6, 0))

        self.unmapped_label = tk.Label(
            inner, text="", font=("Microsoft JhengHei", 9, "bold"), bg="#f5f7fa",
            fg="#dc2626", wraplength=700, justify="left")
        self.unmapped_label.pack(pady=(6, 0))

        btn_frame = tk.Frame(inner, bg="#f5f7fa")
        btn_frame.pack(pady=18)
        self.export_btn = tk.Button(
            btn_frame, text="⬇  导出 Excel", font=("Microsoft JhengHei", 13, "bold"),
            bg="#6366f1", fg="white", activebackground="#4f46e5", activeforeground="white",
            relief="flat", padx=30, pady=10, command=self.on_export, state="disabled", cursor="hand2")
        self.export_btn.pack(side="left", padx=8)
        self.reset_btn = tk.Button(
            btn_frame, text="🔄  重新开始", font=("Microsoft JhengHei", 13),
            bg="white", fg="#6366f1", activebackground="#eef2ff",
            relief="solid", bd=1, padx=24, pady=10, command=self.on_reset, cursor="hand2")
        self.reset_btn.pack(side="left", padx=8)

    # ---------- Tab2：三合一提报 ----------
    def _build_tab2(self):
        inner = self._make_scrollable(self.tab2)

        # 三合一状态
        self.t2_ar_data = None  # 主动提报
        self.t2_rs_data = None  # 费率换签
        self.t2_pm_data = None  # 高端物料
        self.t2_merged_rows = None
        self.t2_warnings = []  # 汇总警告

        zones_frame = tk.Frame(inner, bg="#f5f7fa")
        zones_frame.pack(pady=14)

        self.t2_ar_zone = self._make_drop_zone(
            zones_frame, "主动提报", "#059669", 0, self.on_t2_ar_selected)
        self.t2_rs_zone = self._make_drop_zone(
            zones_frame, "费率换签", "#7c3aed", 1, self.on_t2_rs_selected)
        self.t2_pm_zone = self._make_drop_zone(
            zones_frame, "高端物料", "#dc2626", 2, self.on_t2_pm_selected)

        self.t2_status_label = tk.Label(
            inner, text="请上传任意一个或多个文件以开始",
            font=("Microsoft JhengHei", 11), bg="#f5f7fa", fg="#334155",
            wraplength=700, justify="center")
        self.t2_status_label.pack(pady=(14, 8))

        stats_frame = tk.Frame(inner, bg="#f5f7fa")
        stats_frame.pack(pady=4)
        self.t2_stat_ar = self._make_stat_card(stats_frame, "主动提报行数", 0)
        self.t2_stat_rs = self._make_stat_card(stats_frame, "费率换签行数", 1)
        self.t2_stat_pm = self._make_stat_card(stats_frame, "高端物料行数", 2)
        self.t2_stat_total = self._make_stat_card(stats_frame, "总行数", 3)

        self.t2_warning_label = tk.Label(
            inner, text="", font=("Microsoft JhengHei", 9, "bold"), bg="#f5f7fa",
            fg="#dc2626", wraplength=700, justify="left")
        self.t2_warning_label.pack(pady=(6, 0))

        btn_frame = tk.Frame(inner, bg="#f5f7fa")
        btn_frame.pack(pady=18)
        self.t2_export_btn = tk.Button(
            btn_frame, text="⬇  导出 Excel", font=("Microsoft JhengHei", 13, "bold"),
            bg="#059669", fg="white", activebackground="#047857", activeforeground="white",
            relief="flat", padx=30, pady=10, command=self.on_t2_export, state="disabled", cursor="hand2")
        self.t2_export_btn.pack(side="left", padx=8)
        self.t2_reset_btn = tk.Button(
            btn_frame, text="🔄  重新开始", font=("Microsoft JhengHei", 13),
            bg="white", fg="#059669", activebackground="#ecfdf5",
            relief="solid", bd=1, padx=24, pady=10, command=self.on_t2_reset, cursor="hand2")
        self.t2_reset_btn.pack(side="left", padx=8)

    # ---------- Tab2 文件加载回调 ----------
    def on_t2_ar_selected(self, path, frame, loaded_info, icon):
        try:
            rows, unmapped, cap_warns = parse_active_report(path)
            self.t2_ar_data = {"rows": rows, "unmapped": unmapped, "cap_warnings": cap_warns}
            self._mark_loaded(frame, loaded_info, icon, f"✅ {len(rows)} 行已载入")
            self.t2_try_merge()
        except Exception as e:
            messagebox.showerror("主动提报解析失败", str(e))
            traceback.print_exc()

    def on_t2_rs_selected(self, path, frame, loaded_info, icon):
        try:
            rows, unmapped = parse_rate_sign(path)
            self.t2_rs_data = {"rows": rows, "unmapped": unmapped}
            self._mark_loaded(frame, loaded_info, icon, f"✅ {len(rows)} 行已载入")
            self.t2_try_merge()
        except Exception as e:
            messagebox.showerror("费率换签解析失败", str(e))
            traceback.print_exc()

    def on_t2_pm_selected(self, path, frame, loaded_info, icon):
        try:
            rows, unmapped, cap_warns, oos = parse_premium(path)
            self.t2_pm_data = {"rows": rows, "unmapped": unmapped, "cap_warnings": cap_warns, "out_of_stock": oos}
            self._mark_loaded(frame, loaded_info, icon, f"✅ {len(rows)} 行已载入")
            self.t2_try_merge()
        except Exception as e:
            messagebox.showerror("高端物料解析失败", str(e))
            traceback.print_exc()

    # ---------- Tab2 合并与渲染 ----------
    def t2_try_merge(self):
        if not self.t2_ar_data and not self.t2_rs_data and not self.t2_pm_data:
            self.t2_status_label.configure(text="请上传任意一个或多个文件以开始")
            self.t2_export_btn.configure(state="disabled")
            return

        all_rows = []
        ar_count = rs_count = pm_count = 0
        all_warnings = []

        if self.t2_ar_data:
            all_rows.extend(self.t2_ar_data["rows"])
            ar_count = len(self.t2_ar_data["rows"])
            if self.t2_ar_data["unmapped"]:
                all_warnings.append(f"⚠ 主动提报未映射物料：{'、'.join(self.t2_ar_data['unmapped'])}")
            if self.t2_ar_data.get("cap_warnings"):
                all_warnings.append(f"⚠ 主动提报数量超{CAP_LIMIT}已截断：" + "；".join(self.t2_ar_data["cap_warnings"]))

        if self.t2_rs_data:
            all_rows.extend(self.t2_rs_data["rows"])
            rs_count = len(self.t2_rs_data["rows"])
            if self.t2_rs_data["unmapped"]:
                all_warnings.append(f"⚠ 费率换签未映射物料：{'、'.join(self.t2_rs_data['unmapped'])}")

        if self.t2_pm_data:
            all_rows.extend(self.t2_pm_data["rows"])
            pm_count = len(self.t2_pm_data["rows"])
            if self.t2_pm_data["unmapped"]:
                all_warnings.append(f"⚠ 高端物料未映射物料：{'、'.join(self.t2_pm_data['unmapped'])}")
            if self.t2_pm_data.get("cap_warnings"):
                all_warnings.append(f"⚠ 高端物料数量超{CAP_LIMIT}（请确认）：\n" + "\n".join(self.t2_pm_data["cap_warnings"]))
            if self.t2_pm_data.get("out_of_stock"):
                all_warnings.append(f"⚠ 以下门店选了无货物料（已跳过）：\n" + "\n".join(self.t2_pm_data["out_of_stock"]))

        self.t2_merged_rows = all_rows
        total = len(all_rows)

        parts = []
        if ar_count > 0:
            parts.append(f"主动提报 {ar_count} 行")
        if rs_count > 0:
            parts.append(f"费率换签 {rs_count} 行")
        if pm_count > 0:
            parts.append(f"高端物料 {pm_count} 行")
        msg = f"✅ {'  +  '.join(parts)}，共 {total} 行可导出"
        self.t2_status_label.configure(text=msg)

        self.t2_stat_ar.value_label.configure(text=str(ar_count))
        self.t2_stat_rs.value_label.configure(text=str(rs_count))
        self.t2_stat_pm.value_label.configure(text=str(pm_count))
        self.t2_stat_total.value_label.configure(text=str(total))

        if all_warnings:
            self.t2_warning_label.configure(text="\n".join(all_warnings))
        else:
            self.t2_warning_label.configure(text="")

        self.t2_export_btn.configure(state="normal")

    # ---------- Tab2 导出 ----------
    def on_t2_export(self):
        if not self.t2_merged_rows:
            return
        default_name = f"三合一底板{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        out_dir = get_output_dir()
        save_path = filedialog.asksaveasfilename(
            title="保存三合一底板",
            initialdir=out_dir,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")])
        if not save_path:
            return
        try:
            export_excel(self.t2_merged_rows, save_path)
            messagebox.showinfo("导出成功", f"已保存至：\n{save_path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))
            traceback.print_exc()

    # ---------- Tab2 重置 ----------
    def on_t2_reset(self):
        self.t2_ar_data = None
        self.t2_rs_data = None
        self.t2_pm_data = None
        self.t2_merged_rows = None

        for frame, color in [(self.t2_ar_zone, "#059669"), (self.t2_rs_zone, "#7c3aed"), (self.t2_pm_zone, "#dc2626")]:
            frame.configure(highlightbackground=color, highlightthickness=2)
            frame.icon.configure(fg="#94a3b8")
            frame.loaded_info.configure(text="")

        self.t2_status_label.configure(text="请上传任意一个或多个文件以开始")
        for stat in (self.t2_stat_ar, self.t2_stat_rs, self.t2_stat_pm, self.t2_stat_total):
            stat.value_label.configure(text="0")
        self.t2_warning_label.configure(text="")
        self.t2_export_btn.configure(state="disabled")

    # ---------- 拖拽区组件 ----------
    def _make_drop_zone(self, parent, label, color, col, on_file_selected):
        frame = tk.Frame(parent, width=215, height=150, bg="white",
                         highlightbackground=color, highlightthickness=2, bd=0)
        frame.grid(row=0, column=col, padx=10)
        frame.grid_propagate(False)
        frame.pack_propagate(False)

        tag = tk.Label(frame, text=label, font=("Microsoft JhengHei", 10, "bold"),
                       bg=color, fg="white", padx=8, pady=2)
        tag.place(x=8, y=8)

        icon = tk.Label(frame, text="📁", font=("Segoe UI Emoji", 28), bg="white", fg="#94a3b8")
        icon.place(relx=0.5, rely=0.42, anchor="center")

        hint = tk.Label(frame, text="点击选择 Excel 文件\n(或拖拽到此处)",
                        font=("Microsoft JhengHei", 9), bg="white", fg="#94a3b8", justify="center")
        hint.place(relx=0.5, rely=0.72, anchor="center")

        loaded_info = tk.Label(frame, text="", font=("Microsoft JhengHei", 9, "bold"),
                               bg="white", fg="#059669", wraplength=195, justify="center")
        loaded_info.place(relx=0.5, rely=0.9, anchor="center")

        def choose_file(event=None):
            path = filedialog.askopenfilename(
                title=f"选择 {label} Excel 文件",
                filetypes=[("Excel 文件", "*.xlsx *.xls")])
            if path:
                on_file_selected(path, frame, loaded_info, icon)

        for widget in (frame, icon, hint, tag):
            widget.bind("<Button-1>", choose_file)
            widget.configure(cursor="hand2")

        if HAS_DND:
            frame.drop_target_register(DND_FILES)

            def on_drop(event):
                raw = event.data
                path = raw.strip("{}")
                if path.lower().endswith((".xlsx", ".xls")):
                    on_file_selected(path, frame, loaded_info, icon)
                else:
                    messagebox.showerror("格式错误", "请拖拽 .xlsx / .xls 格式的 Excel 文件")

            frame.dnd_bind("<<Drop>>", on_drop)

        frame.loaded_info = loaded_info
        frame.icon = icon
        return frame

    def _make_stat_card(self, parent, label, col):
        card = tk.Frame(parent, bg="white", highlightbackground="#e2e8f0",
                        highlightthickness=1, padx=18, pady=10)
        card.grid(row=0, column=col, padx=8)
        value_label = tk.Label(card, text="0", font=("Microsoft JhengHei", 22, "bold"),
                               bg="white", fg="#6366f1")
        value_label.pack()
        tk.Label(card, text=label, font=("Microsoft JhengHei", 9), bg="white", fg="#94a3b8").pack()
        card.value_label = value_label
        return card

    def _mark_loaded(self, frame, loaded_info, icon, text):
        frame.configure(highlightbackground="#10b981", highlightthickness=3)
        icon.configure(fg="#10b981")
        loaded_info.configure(text=text)

    # ---------- 文件加载回调 ----------
    def on_qs_selected(self, path, frame, loaded_info, icon):
        try:
            self.qs_data = parse_questionnaire(path)
            self._mark_loaded(frame, loaded_info, icon, f"✅ {len(self.qs_data)} 个门店已载入")
            self.try_merge()
        except Exception as e:
            messagebox.showerror("问卷解析失败", str(e))
            traceback.print_exc()

    def on_sv_selected(self, path, frame, loaded_info, icon):
        try:
            sv_rows, spec_match_count, unmapped_sv = parse_service_order(path)
            self.sv_data = sv_rows
            self.sv_spec_count = spec_match_count
            self.unmapped_sv = unmapped_sv
            matched = sum(1 for v in spec_match_count.values() if v > 0)
            self._mark_loaded(frame, loaded_info, icon,
                              f"✅ {len(sv_rows)} 个门店\n{matched}/{len(SERVICE_CSP_MAPPINGS)} 规格匹配")
            self.try_merge()
        except Exception as e:
            messagebox.showerror("服务订单解析失败", str(e))
            traceback.print_exc()

    def on_vl_selected(self, path, frame, loaded_info, icon):
        try:
            self.vl_data = parse_vlookup(path)
            self._mark_loaded(frame, loaded_info, icon, f"✅ {len(self.vl_data)} 个门店资料已载入")
            self.try_merge()
        except Exception as e:
            messagebox.showerror("门店资料解析失败", str(e))
            traceback.print_exc()

    # ---------- 合并与渲染 ----------
    def try_merge(self):
        if not self.qs_data and not self.sv_data:
            self.status_label.configure(text="请上传「问卷结果」或「商家服务订单」以开始")
            self.export_btn.configure(state="disabled")
            return

        all_rows = []
        if self.qs_data:
            for row in self.qs_data:
                apply_vlookup(row, self.vl_data)
                all_rows.append(row)
        if self.sv_data:
            for row in self.sv_data:
                apply_vlookup(row, self.vl_data)
                all_rows.append(row)

        cap_log = cap_check(all_rows)
        self.merged_rows = all_rows

        qs_count = len(self.qs_data) if self.qs_data else 0
        sv_count = len(self.sv_data) if self.sv_data else 0
        items = count_items(all_rows)
        total = count_total(all_rows)

        if self.qs_data and not self.sv_data:
            msg = f"✅ 问卷结果已载入（{qs_count} 门店），可直接导出，或继续上传商家服务订单"
        elif self.sv_data and not self.qs_data:
            msg = f"✅ 商家服务订单已载入（{sv_count} 门店），可直接导出，或继续上传问卷结果"
        else:
            msg = f"✅ 问卷 {qs_count} 门店 + 商城 {sv_count} 门店，各自独立行输出"
        self.status_label.configure(text=msg)

        self.stat_qs.value_label.configure(text=str(qs_count))
        self.stat_sv.value_label.configure(text=str(sv_count))
        self.stat_items.value_label.configure(text=str(items))
        self.stat_total.value_label.configure(text=str(total))

        if cap_log:
            lines = [f"共 {len(cap_log)} 行有商品数量超过 {CAP_LIMIT}，已截断为 {CAP_LIMIT}："]
            for item in cap_log[:8]:
                lines.append(f"[{item['orderType']}] 门店 {item['shopId']} — {'、'.join(item['details'])}")
            if len(cap_log) > 8:
                lines.append(f"...另有 {len(cap_log) - 8} 行，详见导出文件")
            self.cap_label.configure(text="\n".join(lines))
        else:
            self.cap_label.configure(text="")

        # 未映射规格警告（仅商城）
        if self.unmapped_sv:
            self.unmapped_label.configure(
                text=f"⚠ 商城中以下规格未映射，已跳过：{'、'.join(self.unmapped_sv)}")
        else:
            self.unmapped_label.configure(text="")

        self.export_btn.configure(state="normal")

    # ---------- 导出 ----------
    def on_export(self):
        if not self.merged_rows:
            return
        default_name = f"物料底板{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        out_dir = get_output_dir()
        save_path = filedialog.asksaveasfilename(
            title="保存物料底板",
            initialdir=out_dir,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")])
        if not save_path:
            return
        try:
            export_excel(self.merged_rows, save_path)
            messagebox.showinfo("导出成功", f"已保存至：\n{save_path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))
            traceback.print_exc()

    # ---------- 重置 ----------
    def on_reset(self):
        self.qs_data = None
        self.sv_data = None
        self.sv_spec_count = None
        self.vl_data = None
        self.merged_rows = None
        self.unmapped_sv = []

        for frame in (self.qs_zone, self.sv_zone, self.vl_zone):
            frame.configure(highlightthickness=2)
            frame.icon.configure(fg="#94a3b8")
            frame.loaded_info.configure(text="")
        self.qs_zone.configure(highlightbackground="#6366f1")
        self.sv_zone.configure(highlightbackground="#d97706")
        self.vl_zone.configure(highlightbackground="#16a34a")

        self.status_label.configure(text="请上传「问卷结果」或「商家服务订单」以开始")
        for stat in (self.stat_qs, self.stat_sv, self.stat_items, self.stat_total):
            stat.value_label.configure(text="0")
        self.cap_label.configure(text="")
        self.unmapped_label.configure(text="")
        self.export_btn.configure(state="disabled")


# ============================================================
# 主入口
# ============================================================
def main():
    if HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    app = MaterialToolApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
